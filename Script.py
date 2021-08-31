import json
import openpyxl as pyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from datetime import datetime, date


def add(JSON, Excel):
    month_dict = {
        1: 'January',
        2: 'February',
        3: 'March',
        4: 'April',
        5: 'May ',
        6: 'June ',
        7: 'July',
        8: 'August',
        9: 'September',
        10: 'October',
        11: 'November',
        12: 'December'
    }

    bank_mask = {
        'A': 'ABB',
        'B': 'ABMB',
        'C': 'AGRO',
        'D': 'AMBB',
        'E': 'ARM',
        'F': 'BIMB',
        'G': 'BKRB',
        'H': 'BMMB',
        'I': 'BOC',
        'J': 'BSN',
        'K': 'CIMB',
        'L': 'CITI',
        'M': 'HLB',
        'N': 'HSBC',
        'O': 'KFH',
        'P': 'MBB',
        'Q': 'MBSB',
        'R': 'OCBC',
        'S': 'PBB',
        'T': 'RHB',
        'U': 'SCB',
        'V': 'UOB',
        'W': 'CPAY',
        'X': 'FASP',
        'Y': 'MOB1',
        'Z': 'NETS',
        'A1': 'RMS',
        'A2': 'RSSB'
    }

    loc_dict = {
        "issuer_transaction": {
            'start': {'row': 11, 'col': 6},
            'cmp head': [False],
            'id start': {'row': 11, 'col': 3},
            'id length': 3,
            'order': [
                "contact_volume",
                "contact_value",
                "contactless_volume",
                "contacless_value",
                "sales_volume",
                "sales_value"
            ]
        },
        "acquirer_transaction": {
            'start': {'row': 11, 'col': 6},
            'cmp head': [False],
            'id start': {'row': 11, 'col': 3},
            'id length': 3,
            'order': [
                "contact_volume",
                "contact_value",
                "contactless_volume",
                "contacless_value",
                "sales_volume",
                "sales_value"
            ]
        },
        "denial_codes": {
            'start': {'row': 6, 'col': 5},
            'cmp head': [True, {'row': 5, 'col': 5}],
            'id start': {'row': 6, 'col': 2},
            'id length': 3,
            'order': [
                False
            ]
        },
        'rest': [
            ['title', 'Summary', 'C4'],
            ['title', 'issuer_transaction', 'F4'],
            ['title', 'acquirer_transaction', 'F4'],
        ],
        "system_availability": {
            'Service Maintenance': {
                'start': {'row': 6, 'col': 4},
                'cmp head' : [False],
                'id start': {'row': 6, 'col': 2},
                'id length': 2,
                'order': [
                    'total_maintenance'
                    ],
                'maintenance details': {
                    'start': {'row': 6, 'col': 6}
                    }
                },
                'Issuer Dispute': {
                    'start': {'row': 5, 'col': 4},
                    'cmp head': [False],
                    'id start': {'row': 5, 'col': 2},
                    'id length': 2,
                    'order': [
                        'within_sla',
                        'beyond_sla'
                        ]
                }
            }
        }

    def append_row(start_row, finish_row, col, to_col, ws, to_add=None, only_add=False):
        if not only_add:
            for row in range(finish_row,start_row-1,-1):
                ws.move_range(f'{get_column_letter(col)+str(row)}:{get_column_letter(col+2)+str(row)}', rows=+1, cols=0)
                ws.move_range(f'{get_column_letter(to_col)+str(row)}:{get_column_letter(to_col)+str(row)}', rows=+1, cols=0)
        index = 0
        for col in iter_tool(col, to_col+1, start_row):
            if index == 0:
                to_add[index] = datetime.strptime(to_add[index], '%d/%m/%Y').strftime('%m/%d/%Y')
                ws[col].alignment = Alignment(vertical='center', horizontal='right')
                ws[col].font = Font(size=9)
            elif 0 < index < 3:
                ws[col].alignment = Alignment(vertical='center', horizontal='center')
                ws[col].font = Font(size=10)
            elif index == 4:
                ws[col].alignment = Alignment(vertical='bottom', horizontal='left')
                ws[col].font = Font(size=10)
            ws[col].value = to_add[index]
            ws[col].border = Border(top = Side(border_style='thin', color='FF000000'),
                                          right = Side(border_style='thin', color='FF000000'),
                                          bottom = Side(border_style='thin', color='FF000000'),
                                          left = Side(border_style='thin', color='FF000000'))
            index+= 1

    def iter_tool(start, finish, row):
        for i in range(start, finish):
            yield get_column_letter(i) + str(row)

    def linear_search(sheetname, key, guide):
        WS = WB[sheetname]
        max_row = WS.max_row
        for row in range(guide['start']['row'], max_row):
            address = get_column_letter(guide['start']['col'] - 1) + str(row)
            if WS[address].value in JSON[key] or WS[address].value == JSON['month']:
                the_cell = True
                i = 0
                for col in WS.iter_cols(min_col=guide['id start']['col'],
                                        max_col=guide['id start']['col'] + guide['id length'] - 2,
                                        min_row=row, max_row=row):
                    if str(col[0].value) != JSON[identifier[i]]:
                        the_cell = False
                        break
                    i += 1
                if the_cell and not guide['cmp head'][0]:
                    i = 0
                    for cell in iter_tool(guide['start']['col'],
                                          guide['start']['col'] + len(guide['order']), row):
                        WS[cell].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                        WS[cell].value = guide['order'][i] #for testing purpose, should be JSON[key][guide['order'][i]]
                        i += 1
                elif the_cell:
                    header = iter_tool(guide['start']['col'],
                                       guide['start']['col'] + len(JSON[key][WS[address].value].keys()),
                                       guide['cmp head'][1]['row'])
                    for cell in iter_tool(guide['start']['col'],
                                          guide['start']['col'] + len(JSON[key][WS[address].value].keys()), row):
                        WS[cell].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                        WS[cell].value = WS[next(header)].value #for testing purpose, should be JSON[key][WS[next(header)].value]
            if WS[address].value in bank_mask and bank_mask[WS[address].value] == JSON['bank']:
                WS[address].value = JSON['bank']
                WS[address].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    identifier = ['year', 'month', 'bank']

    JSON = json.load(JSON)
    WB = pyxl.load_workbook(Excel)

    if 'year' not in JSON.keys():
        JSON['year'] = '2021'

    JSON['month'] = month_dict[JSON['month']]
    for key in JSON:
        #for key that act as identifier
        if key in identifier:
            continue
        key_space = key.replace('_', ' ').title()
        #if key isnt implemented yet
        if key_space not in WB.sheetnames and key not in loc_dict:
            print(f'{key_space} not in')
            continue
        #special case - for system availability
        elif key == "system_availability":
            for sheet in loc_dict["system_availability"]:
                linear_search(sheet, key, loc_dict[key][sheet])
                #special case - filling the service maintenance detail
                if sheet == 'Service Maintenance':
                    WS = WB[sheet]
                    start = loc_dict[key][sheet]['maintenance details']['start']['row']
                    collumn = loc_dict[key][sheet]['maintenance details']['start']['col']
                    for data in JSON[key]['datetime']:
                        #setting time from inside data to be comparable
                        the_date = date(*list(map(int, data[0].split('/')))[::-1])
                        the_start = datetime.strptime(data[1], '%I:%M %p').time()
                        #looping through row to find the right place
                        for i in range(start, WS.max_row):
                            colls = list(iter_tool(collumn, collumn+3, i))
                            if not WS[colls[0]].value:
                                add = True
                                break
                            add = True
                            if type(WS[colls[0]].value) == type(''):
                                col_start = datetime.strptime(WS[colls[1]].value, '%I:%M %p').time()
                                col_date = datetime.strptime(WS[colls[0]].value, '%m/%d/%Y').date()
                                if col_date == the_date:
                                    if col_start > the_start:
                                        data.insert(3,'=HOUR(H6-G6)&\" hours, \"&MINUTE(H6-G6)&\" minutes\"'
                                                    .replace('6', str(i)))
                                        append_row(i, WS.max_row, collumn, WS.max_column, WS, data)
                                        add = False
                                        break
                                elif col_date > the_date:
                                    data.insert(3,'=HOUR(H6-G6)&\" hours, \"&MINUTE(H6-G6)&\" minutes\"'
                                                    .replace('6', str(i)))
                                    append_row(i, WS.max_row, collumn, WS.max_column, WS, data)
                                    add = False
                                    break
                            elif WS[colls[0]].value.date() == the_date:
                                if WS[colls[1]].value > the_start:
                                    data.insert(3, '=HOUR(H6-G6)&\" hours, \"&MINUTE(H6-G6)&\" minutes\"'.replace('6',str(i)))
                                    append_row(i, WS.max_row, collumn, WS.max_column, WS, data)
                                    add = False
                                    break
                            elif WS[colls[0]].value.date()>the_date:
                                data.insert(3, '=HOUR(H6-G6)&\" hours, \"&MINUTE(H6-G6)&\" minutes\"'.replace('6',str(i)))
                                append_row(i, WS.max_row, collumn, WS.max_column, WS, data)
                                add = False
                                break
                        if add:
                            data.insert(3, '=HOUR(H6-G6)&\" hours, \"&MINUTE(H6-G6)&\" minutes\"'.replace('6',str(i)))
                            append_row(i, i+1, collumn, WS.max_column, WS, data, True)
                        true_max_row = i
                    row = start
                    while WS[get_column_letter(collumn+2)+str(row)].value:
                        WS[get_column_letter(collumn+3)+str(row)] = '=HOUR(H6-G6)&\" hours, \"&MINUTE(H6-G6)&\" minutes\"'.replace('6',str(row))
                        WS[get_column_letter(collumn+3)+str(row)].border = Border(top = Side(border_style='thin', color='FF000000'),
                                                      right = Side(border_style='thin', color='FF000000'),
                                                      bottom = Side(border_style='thin', color='FF000000'),
                                                      left = Side(border_style='thin', color='FF000000'))
                        WS[get_column_letter(collumn+3)+str(row)].alignment = Alignment(vertical='top', horizontal='center')
                        WS[get_column_letter(collumn+3)+str(row)].font = Font(size=10)
                        row+= 1
        #for the normal case when key written in JSON is the sheet name
        elif key in loc_dict:
            linear_search(key_space, key, loc_dict[key])
    for target, sheet, coor in loc_dict['rest']:
        if target == 'title':
            sheet = sheet.replace('_', ' ').title()
            WS = WB[sheet]
            WS[coor].value = WS[coor].value.replace('ABC', JSON['bank'])
    WB.save(filename=f'{datetime.now().strftime("%m-%d-%Y_%H-%M-%S")}_{JSON[identifier[2]]}.xlsx')


add(open('req.json'), 'excel.xlsx')